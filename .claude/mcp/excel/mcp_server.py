#!/usr/bin/env python3
"""Excel MCP Server - Self-contained openpyxl-based Excel operations."""

import json
import re
from collections import defaultdict
from copy import copy
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
from mcp.server.fastmcp import FastMCP

mcp = FastMCP("excel")

# Resolve project root: .claude/mcp/excel/mcp_server.py -> project root (3 levels up from .claude)
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent
MAX_READ_ROWS = 10000


# ── Helpers ──────────────────────────────────────────────────────────────────

def _resolve_path(file_path: str) -> Path:
    """Resolve file path: absolute passes through, relative resolves from PROJECT_ROOT."""
    p = Path(file_path)
    if p.is_absolute():
        return p
    return PROJECT_ROOT / p


def _open_workbook(file_path: str, data_only: bool = False):
    """Open workbook, returning (wb, resolved_path)."""
    resolved = _resolve_path(file_path)
    if not resolved.exists():
        raise FileNotFoundError(f"File not found: {resolved}")
    if resolved.suffix.lower() == ".xls":
        raise ValueError(".xls format not supported. Convert to .xlsx first.")
    wb = openpyxl.load_workbook(str(resolved), data_only=data_only)
    return wb, resolved


def _get_sheet(wb, sheet_name: Optional[str] = None):
    """Get worksheet by name or return active sheet."""
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        return wb[sheet_name]
    return wb.active


def _serialize(value):
    """Convert cell value to JSON-serializable type."""
    if value is None:
        return None
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)


def _ok(**kwargs):
    return json.dumps({"success": True, **kwargs})


def _err(msg: str):
    return json.dumps({"success": False, "error": msg})


def _validate_range_str(range_str: str) -> bool:
    """Check if string like 'A1:C10' or 'A1' is a valid Excel range."""
    return bool(re.match(r'^[A-Za-z]{1,3}\d+(?::[A-Za-z]{1,3}\d+)?$', range_str))


# ── GROUP A: Core Read/Write ─────────────────────────────────────────────────

@mcp.tool()
def create_workbook(file_path: str, sheet_name: str = "Sheet1") -> str:
    """Create a new .xlsx workbook file.

    Args:
        file_path: Path for new file (absolute or relative to project root)
        sheet_name: Name for the default sheet
    """
    try:
        resolved = _resolve_path(file_path)
        resolved.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        wb.active.title = sheet_name
        wb.save(str(resolved))
        return _ok(file_path=str(resolved), sheet=sheet_name)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def get_workbook_metadata(file_path: str) -> str:
    """Get workbook metadata: sheets, dimensions, row/column counts.

    Args:
        file_path: Path to .xlsx file
    """
    try:
        wb, resolved = _open_workbook(file_path)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets.append({
                "name": name,
                "dimensions": ws.dimensions,
                "min_row": ws.min_row,
                "max_row": ws.max_row,
                "min_column": ws.min_column,
                "max_column": ws.max_column,
                "rows": (ws.max_row or 0) - (ws.min_row or 0) + 1 if ws.max_row else 0,
                "columns": (ws.max_column or 0) - (ws.min_column or 0) + 1 if ws.max_column else 0,
            })
        return _ok(file_path=str(resolved), sheets=sheets)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def write_data_to_excel(
    file_path: str,
    data: list,
    sheet_name: str = None,
    start_cell: str = "A1",
    create_if_missing: bool = True,
) -> str:
    """Write data (list of lists) to an Excel worksheet.

    Args:
        file_path: Path to .xlsx file
        data: List of rows, each row is a list of values. Example: [["Name","Age"],["Alice",30]]
        sheet_name: Target sheet (default: active sheet)
        start_cell: Cell to start writing from (e.g. "A1")
        create_if_missing: Create file if it doesn't exist
    """
    try:
        resolved = _resolve_path(file_path)
        if not resolved.exists():
            if create_if_missing:
                resolved.parent.mkdir(parents=True, exist_ok=True)
                wb = openpyxl.Workbook()
            else:
                raise FileNotFoundError(f"File not found: {resolved}")
        else:
            wb = openpyxl.load_workbook(str(resolved))

        ws = _get_sheet(wb, sheet_name)
        min_col, min_row, _, _ = range_boundaries(f"{start_cell}:{start_cell}")
        cells_written = 0

        for r_idx, row in enumerate(data):
            for c_idx, value in enumerate(row):
                ws.cell(row=min_row + r_idx, column=min_col + c_idx, value=value)
                cells_written += 1

        end_row = min_row + len(data) - 1
        end_col = min_col + max(len(r) for r in data) - 1 if data else min_col
        end_cell = f"{get_column_letter(end_col)}{end_row}"

        wb.save(str(resolved))
        return _ok(cells_written=cells_written, range=f"{start_cell}:{end_cell}")
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def read_data_from_excel(
    file_path: str,
    sheet_name: str = None,
    range: str = None,
    max_rows: int = 1000,
    headers: bool = True,
) -> str:
    """Read data from an Excel worksheet.

    Args:
        file_path: Path to .xlsx file
        sheet_name: Sheet to read (default: active sheet)
        range: Cell range to read (e.g. "A1:D50"). If omitted, reads entire used range
        max_rows: Maximum rows to return (safety limit)
        headers: If true, first row becomes dict keys; otherwise returns list of lists
    """
    try:
        wb, resolved = _open_workbook(file_path, data_only=True)
        ws = _get_sheet(wb, sheet_name)

        if range and _validate_range_str(range):
            min_c, min_r, max_c, max_r = range_boundaries(range)
        else:
            min_r = ws.min_row or 1
            min_c = ws.min_column or 1
            max_r = ws.max_row or 1
            max_c = ws.max_column or 1

        total_rows = max_r - min_r + 1
        cap = min(max_rows, MAX_READ_ROWS)
        actual_max_r = min(max_r, min_r + cap - 1)

        rows = []
        for row in ws.iter_rows(min_row=min_r, max_row=actual_max_r,
                                min_col=min_c, max_col=max_c):
            rows.append([_serialize(cell.value) for cell in row])

        if headers and len(rows) > 1:
            header_row = rows[0]
            data = []
            for row in rows[1:]:
                data.append(dict(zip(header_row, row)))
            return _ok(data=data, rows_read=len(data), total_rows=total_rows - 1,
                       sheet=ws.title, headers=header_row)
        else:
            return _ok(data=rows, rows_read=len(rows), total_rows=total_rows,
                       sheet=ws.title)
    except Exception as e:
        return _err(str(e))


# ── GROUP B: Worksheet Management ────────────────────────────────────────────

@mcp.tool()
def create_worksheet(
    file_path: str,
    sheet_name: str,
    position: int = None,
) -> str:
    """Add a new worksheet to an existing workbook.

    Args:
        file_path: Path to .xlsx file
        sheet_name: Name for the new sheet
        position: 0-indexed position to insert at (default: append at end)
    """
    try:
        wb, resolved = _open_workbook(file_path)
        if sheet_name in wb.sheetnames:
            return _err(f"Sheet '{sheet_name}' already exists")
        wb.create_sheet(title=sheet_name, index=position)
        wb.save(str(resolved))
        return _ok(sheet=sheet_name, sheets=wb.sheetnames)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def rename_worksheet(file_path: str, old_name: str, new_name: str) -> str:
    """Rename a worksheet.

    Args:
        file_path: Path to .xlsx file
        old_name: Current sheet name
        new_name: New sheet name
    """
    try:
        wb, resolved = _open_workbook(file_path)
        ws = _get_sheet(wb, old_name)
        ws.title = new_name
        wb.save(str(resolved))
        return _ok(old_name=old_name, new_name=new_name, sheets=wb.sheetnames)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def copy_worksheet(
    file_path: str,
    source_sheet: str,
    new_name: str = None,
) -> str:
    """Duplicate an existing worksheet.

    Args:
        file_path: Path to .xlsx file
        source_sheet: Name of sheet to copy
        new_name: Name for the copy (default: auto-generated)
    """
    try:
        wb, resolved = _open_workbook(file_path)
        ws = _get_sheet(wb, source_sheet)
        new_ws = wb.copy_worksheet(ws)
        if new_name:
            new_ws.title = new_name
        wb.save(str(resolved))
        return _ok(source=source_sheet, new_sheet=new_ws.title, sheets=wb.sheetnames)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def delete_worksheet(file_path: str, sheet_name: str) -> str:
    """Delete a worksheet from the workbook.

    Args:
        file_path: Path to .xlsx file
        sheet_name: Name of sheet to delete
    """
    try:
        wb, resolved = _open_workbook(file_path)
        if len(wb.sheetnames) <= 1:
            return _err("Cannot delete the last sheet in a workbook")
        ws = _get_sheet(wb, sheet_name)
        wb.remove(ws)
        wb.save(str(resolved))
        return _ok(deleted=sheet_name, sheets=wb.sheetnames)
    except Exception as e:
        return _err(str(e))


# ── GROUP C: Cell/Range Operations ───────────────────────────────────────────

@mcp.tool()
def apply_formula(
    file_path: str,
    cell: str,
    formula: str,
    sheet_name: str = None,
) -> str:
    """Set an Excel formula in a cell.

    Args:
        file_path: Path to .xlsx file
        cell: Cell reference (e.g. "A1")
        formula: Formula string (e.g. "=SUM(B1:B10)"). Prepends = if missing.
        sheet_name: Target sheet (default: active)
    """
    try:
        wb, resolved = _open_workbook(file_path)
        ws = _get_sheet(wb, sheet_name)
        if not formula.startswith("="):
            formula = "=" + formula
        ws[cell] = formula
        wb.save(str(resolved))
        return _ok(cell=cell, formula=formula, sheet=ws.title)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def validate_formula_syntax(formula: str) -> str:
    """Check if an Excel formula is syntactically plausible.

    Args:
        formula: Formula string to validate (e.g. "=SUM(A1:A10)")
    """
    issues = []
    if not formula.startswith("="):
        issues.append("Formula must start with '='")

    body = formula[1:] if formula.startswith("=") else formula
    if body.count("(") != body.count(")"):
        issues.append(f"Unbalanced parentheses: {body.count('(')} open, {body.count(')')} close")

    if re.search(r'[+\-*/^&=<>],', body):
        issues.append("Operator followed by comma")
    if re.search(r',[+\-*/^&=<>]', body):
        issues.append("Comma followed by operator")
    if body.endswith(tuple("+-*/^&=<>")):
        issues.append("Formula ends with operator")

    return json.dumps({"valid": len(issues) == 0, "formula": formula, "issues": issues})


@mcp.tool()
def format_range(
    file_path: str,
    range: str,
    sheet_name: str = None,
    bold: bool = None,
    italic: bool = None,
    font_size: int = None,
    font_color: str = None,
    bg_color: str = None,
    number_format: str = None,
    alignment: str = None,
    border: str = None,
) -> str:
    """Apply formatting to a cell range.

    Args:
        file_path: Path to .xlsx file
        range: Cell range (e.g. "A1:D10")
        sheet_name: Target sheet (default: active)
        bold: Make text bold
        italic: Make text italic
        font_size: Font size in points
        font_color: Font color hex (e.g. "FF0000" for red)
        bg_color: Background color hex (e.g. "FFFF00" for yellow)
        number_format: Excel number format (e.g. "#,##0.00", "0%", "yyyy-mm-dd")
        alignment: Text alignment - "left", "center", "right"
        border: Border style - "thin", "medium", "thick"
    """
    try:
        wb, resolved = _open_workbook(file_path)
        ws = _get_sheet(wb, sheet_name)
        min_c, min_r, max_c, max_r = range_boundaries(range)
        count = 0

        for row in ws.iter_rows(min_row=min_r, max_row=max_r,
                                min_col=min_c, max_col=max_c):
            for cell in row:
                # Font
                font_kwargs = {}
                if cell.font:
                    font_kwargs = {
                        "name": cell.font.name,
                        "size": cell.font.size,
                        "bold": cell.font.bold,
                        "italic": cell.font.italic,
                        "color": cell.font.color,
                    }
                if bold is not None:
                    font_kwargs["bold"] = bold
                if italic is not None:
                    font_kwargs["italic"] = italic
                if font_size is not None:
                    font_kwargs["size"] = font_size
                if font_color is not None:
                    font_kwargs["color"] = font_color
                if font_kwargs:
                    cell.font = Font(**font_kwargs)

                # Fill
                if bg_color is not None:
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

                # Number format
                if number_format is not None:
                    cell.number_format = number_format

                # Alignment
                if alignment is not None:
                    horiz = alignment if alignment in ("left", "center", "right") else None
                    vert = alignment if alignment in ("top", "middle", "bottom") else None
                    if vert == "middle":
                        vert = "center"
                    cell.alignment = Alignment(horizontal=horiz, vertical=vert)

                # Border
                if border is not None:
                    side = Side(style=border)
                    cell.border = Border(left=side, right=side, top=side, bottom=side)

                count += 1

        wb.save(str(resolved))
        return _ok(cells_formatted=count, range=range, sheet=ws.title)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def merge_cells(file_path: str, range: str, sheet_name: str = None) -> str:
    """Merge a range of cells.

    Args:
        file_path: Path to .xlsx file
        range: Cell range to merge (e.g. "A1:D1")
        sheet_name: Target sheet (default: active)
    """
    try:
        wb, resolved = _open_workbook(file_path)
        ws = _get_sheet(wb, sheet_name)
        ws.merge_cells(range)
        wb.save(str(resolved))
        return _ok(merged=range, sheet=ws.title)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def unmerge_cells(file_path: str, range: str, sheet_name: str = None) -> str:
    """Unmerge a previously merged cell range.

    Args:
        file_path: Path to .xlsx file
        range: Cell range to unmerge (e.g. "A1:D1")
        sheet_name: Target sheet (default: active)
    """
    try:
        wb, resolved = _open_workbook(file_path)
        ws = _get_sheet(wb, sheet_name)
        ws.unmerge_cells(range)
        wb.save(str(resolved))
        return _ok(unmerged=range, sheet=ws.title)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def get_merged_cells(file_path: str, sheet_name: str = None) -> str:
    """List all merged cell ranges in a worksheet.

    Args:
        file_path: Path to .xlsx file
        sheet_name: Target sheet (default: active)
    """
    try:
        wb, resolved = _open_workbook(file_path)
        ws = _get_sheet(wb, sheet_name)
        merged = [str(r) for r in ws.merged_cells.ranges]
        return _ok(merged_ranges=merged, count=len(merged), sheet=ws.title)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def copy_range(
    file_path: str,
    source_range: str,
    target_start: str,
    sheet_name: str = None,
    target_sheet: str = None,
) -> str:
    """Copy cell values from one range to another.

    Args:
        file_path: Path to .xlsx file
        source_range: Source range (e.g. "A1:C10")
        target_start: Top-left cell for paste (e.g. "E1")
        sheet_name: Source sheet (default: active)
        target_sheet: Destination sheet (default: same as source)
    """
    try:
        wb, resolved = _open_workbook(file_path)
        src_ws = _get_sheet(wb, sheet_name)
        tgt_ws = _get_sheet(wb, target_sheet) if target_sheet else src_ws

        s_min_c, s_min_r, s_max_c, s_max_r = range_boundaries(source_range)
        t_min_c, t_min_r, _, _ = range_boundaries(f"{target_start}:{target_start}")

        cells_copied = 0
        for r_offset in range(s_max_r - s_min_r + 1):
            for c_offset in range(s_max_c - s_min_c + 1):
                src_val = src_ws.cell(row=s_min_r + r_offset, column=s_min_c + c_offset).value
                tgt_ws.cell(row=t_min_r + r_offset, column=t_min_c + c_offset, value=src_val)
                cells_copied += 1

        wb.save(str(resolved))
        return _ok(cells_copied=cells_copied, source=source_range, target_start=target_start)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def delete_range(file_path: str, range: str, sheet_name: str = None) -> str:
    """Clear contents of a cell range (sets cells to None).

    Args:
        file_path: Path to .xlsx file
        range: Cell range to clear (e.g. "A1:C10")
        sheet_name: Target sheet (default: active)
    """
    try:
        wb, resolved = _open_workbook(file_path)
        ws = _get_sheet(wb, sheet_name)
        min_c, min_r, max_c, max_r = range_boundaries(range)
        count = 0

        for row in ws.iter_rows(min_row=min_r, max_row=max_r,
                                min_col=min_c, max_col=max_c):
            for cell in row:
                cell.value = None
                count += 1

        wb.save(str(resolved))
        return _ok(cells_cleared=count, range=range, sheet=ws.title)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def validate_excel_range(
    range: str,
    file_path: str = None,
    sheet_name: str = None,
) -> str:
    """Check if a range string is a valid Excel range reference.

    Args:
        range: Range to validate (e.g. "A1:Z100" or "A1")
        file_path: Optional - if provided, also checks range is within sheet dimensions
        sheet_name: Sheet to check against (default: active)
    """
    issues = []
    if not _validate_range_str(range):
        issues.append(f"Invalid range format: '{range}'. Expected pattern like 'A1' or 'A1:C10'")

    if not issues and file_path:
        try:
            wb, _ = _open_workbook(file_path)
            ws = _get_sheet(wb, sheet_name)
            min_c, min_r, max_c, max_r = range_boundaries(range)
            if ws.max_row and max_r > ws.max_row:
                issues.append(f"Row {max_r} exceeds sheet max row {ws.max_row}")
            if ws.max_column and max_c > ws.max_column:
                issues.append(f"Column {max_c} exceeds sheet max column {ws.max_column}")
        except Exception as e:
            issues.append(str(e))

    return json.dumps({"valid": len(issues) == 0, "range": range, "issues": issues})


# ── GROUP D: Row/Column Operations ───────────────────────────────────────────

@mcp.tool()
def insert_rows(
    file_path: str,
    row: int,
    count: int = 1,
    sheet_name: str = None,
) -> str:
    """Insert empty rows at a position.

    Args:
        file_path: Path to .xlsx file
        row: Row number to insert before (1-indexed)
        count: Number of rows to insert
        sheet_name: Target sheet (default: active)
    """
    try:
        wb, resolved = _open_workbook(file_path)
        ws = _get_sheet(wb, sheet_name)
        ws.insert_rows(row, count)
        wb.save(str(resolved))
        return _ok(inserted_at=row, count=count, sheet=ws.title)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def insert_columns(
    file_path: str,
    column: int,
    count: int = 1,
    sheet_name: str = None,
) -> str:
    """Insert empty columns at a position.

    Args:
        file_path: Path to .xlsx file
        column: Column number to insert before (1-indexed, A=1)
        count: Number of columns to insert
        sheet_name: Target sheet (default: active)
    """
    try:
        wb, resolved = _open_workbook(file_path)
        ws = _get_sheet(wb, sheet_name)
        ws.insert_cols(column, count)
        wb.save(str(resolved))
        return _ok(inserted_at=column, count=count, sheet=ws.title)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def delete_sheet_rows(
    file_path: str,
    row: int,
    count: int = 1,
    sheet_name: str = None,
) -> str:
    """Delete rows from a worksheet.

    Args:
        file_path: Path to .xlsx file
        row: Starting row number to delete (1-indexed)
        count: Number of rows to delete
        sheet_name: Target sheet (default: active)
    """
    try:
        wb, resolved = _open_workbook(file_path)
        ws = _get_sheet(wb, sheet_name)
        ws.delete_rows(row, count)
        wb.save(str(resolved))
        return _ok(deleted_from=row, count=count, sheet=ws.title)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def delete_sheet_columns(
    file_path: str,
    column: int,
    count: int = 1,
    sheet_name: str = None,
) -> str:
    """Delete columns from a worksheet.

    Args:
        file_path: Path to .xlsx file
        column: Starting column number to delete (1-indexed, A=1)
        count: Number of columns to delete
        sheet_name: Target sheet (default: active)
    """
    try:
        wb, resolved = _open_workbook(file_path)
        ws = _get_sheet(wb, sheet_name)
        ws.delete_cols(column, count)
        wb.save(str(resolved))
        return _ok(deleted_from=column, count=count, sheet=ws.title)
    except Exception as e:
        return _err(str(e))


# ── GROUP E: Advanced Features ───────────────────────────────────────────────

@mcp.tool()
def create_chart(
    file_path: str,
    chart_type: str,
    data_range: str,
    title: str = "",
    sheet_name: str = None,
    categories_range: str = None,
    target_cell: str = "E1",
    width: int = 15,
    height: int = 10,
) -> str:
    """Create a chart in the worksheet.

    Args:
        file_path: Path to .xlsx file
        chart_type: "bar", "line", "pie", or "scatter"
        data_range: Range containing chart data (e.g. "B1:B10")
        title: Chart title
        sheet_name: Sheet containing data (default: active)
        categories_range: Range for category labels (e.g. "A1:A10")
        target_cell: Cell to place the chart at (e.g. "E1")
        width: Chart width in cm (default 15)
        height: Chart height in cm (default 10)
    """
    try:
        wb, resolved = _open_workbook(file_path)
        ws = _get_sheet(wb, sheet_name)

        chart_classes = {
            "bar": BarChart,
            "line": LineChart,
            "pie": PieChart,
            "scatter": ScatterChart,
        }
        if chart_type not in chart_classes:
            return _err(f"Unsupported chart type '{chart_type}'. Use: {list(chart_classes.keys())}")

        chart = chart_classes[chart_type]()
        chart.title = title
        chart.width = width
        chart.height = height

        d_min_c, d_min_r, d_max_c, d_max_r = range_boundaries(data_range)
        data_ref = Reference(ws, min_col=d_min_c, min_row=d_min_r,
                             max_col=d_max_c, max_row=d_max_r)

        if chart_type == "scatter":
            from openpyxl.chart import Series
            if categories_range:
                cat_min_c, cat_min_r, cat_max_c, cat_max_r = range_boundaries(categories_range)
                x_ref = Reference(ws, min_col=cat_min_c, min_row=cat_min_r,
                                  max_col=cat_max_c, max_row=cat_max_r)
                series = Series(data_ref, xvalues=x_ref, title=title)
            else:
                series = Series(data_ref, title=title)
            chart.series.append(series)
        else:
            chart.add_data(data_ref, titles_from_data=True)
            if categories_range:
                cat_min_c, cat_min_r, cat_max_c, cat_max_r = range_boundaries(categories_range)
                cats = Reference(ws, min_col=cat_min_c, min_row=cat_min_r,
                                 max_col=cat_max_c, max_row=cat_max_r)
                chart.set_categories(cats)

        ws.add_chart(chart, target_cell)
        wb.save(str(resolved))
        return _ok(chart_type=chart_type, data_range=data_range, target_cell=target_cell,
                    sheet=ws.title)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def create_pivot_table(
    file_path: str,
    source_range: str,
    target_cell: str,
    rows: list,
    values: list,
    agg_func: str = "sum",
    sheet_name: str = None,
    target_sheet: str = None,
) -> str:
    """Create a pivot table (Python-computed, written as static data).

    Note: openpyxl cannot create native Excel pivot tables. This reads data,
    groups by row fields, aggregates value fields, and writes the result.

    Args:
        file_path: Path to .xlsx file
        source_range: Range with headers in first row (e.g. "A1:D100")
        target_cell: Where to write the pivot result (e.g. "F1")
        rows: Column names to group by (e.g. ["Category", "Region"])
        values: Column names to aggregate (e.g. ["Sales", "Quantity"])
        agg_func: Aggregation function - "sum", "count", "average", "min", "max"
        sheet_name: Source data sheet (default: active)
        target_sheet: Output sheet (default: same as source, or creates "Pivot" sheet)
    """
    try:
        wb, resolved = _open_workbook(file_path, data_only=True)
        ws = _get_sheet(wb, sheet_name)

        # Read source data
        min_c, min_r, max_c, max_r = range_boundaries(source_range)
        all_rows = []
        for row in ws.iter_rows(min_row=min_r, max_row=max_r,
                                min_col=min_c, max_col=max_c, values_only=True):
            all_rows.append(list(row))

        if len(all_rows) < 2:
            return _err("Source range must have a header row and at least one data row")

        headers = [str(h) for h in all_rows[0]]
        data_rows = all_rows[1:]

        # Validate column names
        for col in rows + values:
            if col not in headers:
                return _err(f"Column '{col}' not found. Available: {headers}")

        row_indices = [headers.index(c) for c in rows]
        val_indices = [headers.index(c) for c in values]

        # Group and aggregate
        groups = defaultdict(lambda: defaultdict(list))
        for dr in data_rows:
            key = tuple(str(dr[i]) for i in row_indices)
            for vi in val_indices:
                val = dr[vi]
                if val is not None:
                    try:
                        groups[key][headers[vi]].append(float(val))
                    except (ValueError, TypeError):
                        if agg_func == "count":
                            groups[key][headers[vi]].append(1)

        agg_funcs = {
            "sum": lambda vals: sum(vals),
            "count": lambda vals: len(vals),
            "average": lambda vals: sum(vals) / len(vals) if vals else 0,
            "min": lambda vals: min(vals) if vals else 0,
            "max": lambda vals: max(vals) if vals else 0,
        }
        if agg_func not in agg_funcs:
            return _err(f"Unknown agg_func '{agg_func}'. Use: {list(agg_funcs.keys())}")

        fn = agg_funcs[agg_func]

        # Build result table
        pivot_headers = rows + [f"{v} ({agg_func})" for v in values]
        pivot_data = [pivot_headers]
        for key in sorted(groups.keys()):
            row_data = list(key)
            for v in values:
                vals = groups[key].get(v, [])
                row_data.append(fn(vals) if vals else 0)
            pivot_data.append(row_data)

        # Write to target
        if target_sheet:
            if target_sheet not in wb.sheetnames:
                wb.create_sheet(title=target_sheet)
            tgt_ws = wb[target_sheet]
        else:
            tgt_ws = ws

        t_min_c, t_min_r, _, _ = range_boundaries(f"{target_cell}:{target_cell}")
        for r_idx, prow in enumerate(pivot_data):
            for c_idx, val in enumerate(prow):
                tgt_ws.cell(row=t_min_r + r_idx, column=t_min_c + c_idx, value=val)

        wb.save(str(resolved))
        return _ok(
            rows_in_pivot=len(pivot_data) - 1,
            columns=pivot_headers,
            target_cell=target_cell,
            target_sheet=tgt_ws.title,
        )
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def create_table(
    file_path: str,
    range: str,
    table_name: str,
    style: str = "TableStyleMedium2",
    sheet_name: str = None,
) -> str:
    """Create a formatted Excel Table.

    Args:
        file_path: Path to .xlsx file
        range: Data range including headers (e.g. "A1:D10")
        table_name: Table display name (must be unique in workbook)
        style: Table style name (default: "TableStyleMedium2")
        sheet_name: Target sheet (default: active)
    """
    try:
        wb, resolved = _open_workbook(file_path)
        ws = _get_sheet(wb, sheet_name)

        tab = Table(displayName=table_name, ref=range)
        tab.tableStyleInfo = TableStyleInfo(
            name=style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)
        wb.save(str(resolved))
        return _ok(table_name=table_name, range=range, style=style, sheet=ws.title)
    except Exception as e:
        return _err(str(e))


@mcp.tool()
def get_data_validation_info(file_path: str, sheet_name: str = None) -> str:
    """Get data validation rules from a worksheet.

    Args:
        file_path: Path to .xlsx file
        sheet_name: Target sheet (default: active)
    """
    try:
        wb, resolved = _open_workbook(file_path)
        ws = _get_sheet(wb, sheet_name)
        validations = []
        for dv in ws.data_validations.dataValidation:
            validations.append({
                "type": dv.type,
                "formula1": str(dv.formula1) if dv.formula1 else None,
                "formula2": str(dv.formula2) if dv.formula2 else None,
                "cells": str(dv.sqref),
                "operator": dv.operator,
                "allow_blank": dv.allow_blank,
                "error_message": dv.error,
                "prompt_message": dv.prompt,
            })
        return _ok(validations=validations, count=len(validations), sheet=ws.title)
    except Exception as e:
        return _err(str(e))


# ── Entry Point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    mcp.run()
